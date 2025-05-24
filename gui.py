from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QLabel,
    QPushButton,
    QTabWidget,
    QFormLayout,
    QLineEdit,
    QMessageBox,
    QGroupBox,
    QStyle,
    QCheckBox,
)
from PyQt5.QtGui import QFont, QIcon, QDoubleValidator
from PyQt5.QtCore import Qt
import sys
import openpyxl
from datetime import datetime
import logging
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from data_templates import ozet_roi_data, maliyet_tasarrufu_data, verimlilik_artisi_data, kalite_iyilestirme_data, npv_roi_bilgi_data
from calculations import (
    maliyet_tasarrufu_hesapla,
    verimlilik_artisi_hesapla,
    kalite_iyilestirme_hesapla,
)
from xlsx_report import create_xlsxwriter_report
from formatting import get_common_styles

logging.basicConfig(level=logging.ERROR)

class ROIHesaplamaArayuzu(QMainWindow):
    """Main window providing forms to collect ROI parameters."""

    def __init__(self):
        """Initialize the window and build the interface."""
        super().__init__()
        self.initUI()

    def initUI(self):
        """Create widgets and layout for all tabs."""
        self.setWindowTitle('ROI Hesaplama Aracı')
        self.setGeometry(100, 100, 800, 600)

        # Merkezi widget
        merkezi_widget = QWidget()
        self.setCentralWidget(merkezi_widget)

        # Ana düzen
        layout = QVBoxLayout()
        merkezi_widget.setLayout(layout)

        # Başlık
        baslik = QLabel('Otomasyon ROI Hesaplama Aracı')
        baslik.setFont(QFont('Calibri', 16, QFont.Bold))
        baslik.setAlignment(Qt.AlignCenter)
        layout.addWidget(baslik)

        # Tab Widget Oluştur
        self.tab_widget = QTabWidget()
        layout.addWidget(self.tab_widget)


        # Sayfaları Oluştur
        self.sirket_bilgileri_sayfasi()
        self.maliyet_tasarrufu_sayfasi()
        self.verimlilik_sayfasi()
        self.kalite_sayfasi()

        # Gelişmiş rapor seçeneği
        self.use_xlsxwriter = QCheckBox("Gelişmiş Excel Raporu (xlsxwriter)")
        self.use_xlsxwriter.setChecked(True)
        self.use_xlsxwriter.setToolTip(
            "İşaretli bırakıldığında grafik_ekle_xlsxwriter fonksiyonundaki dört grafikli rapor oluşturulur."
        )
        layout.addWidget(self.use_xlsxwriter)

        # Hesapla Butonu
        hesapla_btn = QPushButton('ROI Hesapla')
        hesapla_btn.setObjectName('calculateButton')
        hesapla_btn.clicked.connect(self.roi_hesapla)
        layout.addWidget(hesapla_btn)

        # Stil
        self.setStyleSheet(
            """
            QMainWindow { background-color: #F0F0F0; }
            QLabel { color: #4F81BD; padding: 5px; }
            QGroupBox { border: 1px solid #4F81BD; border-radius: 5px; margin-top: 10px; }
            QGroupBox::title { subcontrol-origin: margin; subcontrol-position: top left; padding: 0 3px; color: #4F81BD; font-weight: bold; }
            QLineEdit { background: #FFFFFF; padding: 3px; }
            QPushButton#calculateButton { background-color: #4F81BD; color: white; padding: 8px 12px; font-weight: bold; }
        """
        )

    def sirket_bilgileri_sayfasi(self):
        """Create the tab for entering company information."""
        sayfa = QWidget()
        ana_layout = QVBoxLayout()
        sayfa.setLayout(ana_layout)

        grup = QGroupBox("Şirket Bilgileri")
        layout = QFormLayout()

        # Şirket bilgileri input alanları
        self.sirket_adi = QLineEdit()
        self.proje_adi = QLineEdit()
        self.yetkili_kisi = QLineEdit()

        layout.addRow("Şirket Adı:", self.sirket_adi)
        layout.addRow("Proje Adı:", self.proje_adi)
        layout.addRow("Yetkili Kişi:", self.yetkili_kisi)

        grup.setLayout(layout)
        ana_layout.addWidget(grup)

        icon = self.style().standardIcon(QStyle.SP_FileDialogInfoView)
        self.tab_widget.addTab(sayfa, icon, "Şirket Bilgileri")

    def maliyet_tasarrufu_sayfasi(self):
        """Create the tab for labor cost comparison."""
        sayfa = QWidget()
        ana_layout = QVBoxLayout()
        sayfa.setLayout(ana_layout)

        mevcut_grup = QGroupBox("Mevcut Durum")
        mevcut_layout = QFormLayout()

        # Mevcut Durum Input Alanları
        self.mevcut_isci_sayisi = QLineEdit()
        self.mevcut_isci_sayisi.setValidator(QDoubleValidator(0.0, 1e12, 2))
        self.ortalama_maas = QLineEdit()
        self.ortalama_maas.setValidator(QDoubleValidator(0.0, 1e12, 2))
        self.mevcut_vardiya_sayisi = QLineEdit()
        self.mevcut_vardiya_sayisi.setValidator(QDoubleValidator(0.0, 1e12, 2))

        # Otomasyon Sonrası Input Alanları
        self.otomasyon_sonrasi_isci_sayisi = QLineEdit()
        self.otomasyon_sonrasi_isci_sayisi.setValidator(QDoubleValidator(0.0, 1e12, 2))
        self.otomasyon_sonrasi_maas = QLineEdit()
        self.otomasyon_sonrasi_maas.setValidator(QDoubleValidator(0.0, 1e12, 2))
        self.otomasyon_sonrasi_vardiya_sayisi = QLineEdit()
        self.otomasyon_sonrasi_vardiya_sayisi.setValidator(QDoubleValidator(0.0, 1e12, 2))

        # Mevcut Durum Grubu
        mevcut_layout.addRow("Mevcut İşçi Sayısı:", self.mevcut_isci_sayisi)
        mevcut_layout.addRow("Ortalama Aylık Maaş:", self.ortalama_maas)
        mevcut_layout.addRow("Vardiya Sayısı:", self.mevcut_vardiya_sayisi)
        mevcut_grup.setLayout(mevcut_layout)

        otomasyon_grup = QGroupBox("Otomasyon Sonrası")
        otomasyon_layout = QFormLayout()
        otomasyon_layout.addRow("İşçi Sayısı:", self.otomasyon_sonrasi_isci_sayisi)
        otomasyon_layout.addRow("Ortalama Aylık Maaş:", self.otomasyon_sonrasi_maas)
        otomasyon_layout.addRow("Vardiya Sayısı:", self.otomasyon_sonrasi_vardiya_sayisi)
        otomasyon_grup.setLayout(otomasyon_layout)

        ana_layout.addWidget(mevcut_grup)
        ana_layout.addWidget(otomasyon_grup)

        icon = self.style().standardIcon(QStyle.SP_DriveHDIcon)
        self.tab_widget.addTab(sayfa, icon, "Maliyet Tasarrufu")

    def verimlilik_sayfasi(self):
        """Create the tab for productivity inputs."""
        sayfa = QWidget()
        ana_layout = QVBoxLayout()
        sayfa.setLayout(ana_layout)

        mevcut_grup = QGroupBox("Mevcut Durum")
        mevcut_layout = QFormLayout()

        # Mevcut Durum Input Alanları
        self.max_kapasite = QLineEdit()
        self.max_kapasite.setValidator(QDoubleValidator(0.0, 1e12, 2))
        self.oee_mevcut = QLineEdit()
        self.oee_mevcut.setValidator(QDoubleValidator(0.0, 100.0, 2))
        self.calisma_gunu = QLineEdit()
        self.calisma_gunu.setValidator(QDoubleValidator(0.0, 366.0, 0))

        # Otomasyon Sonrası Input Alanları
        self.otomasyon_sonrasi_max_kapasite = QLineEdit()
        self.otomasyon_sonrasi_max_kapasite.setValidator(QDoubleValidator(0.0, 1e12, 2))
        self.otomasyon_sonrasi_oee = QLineEdit()
        self.otomasyon_sonrasi_oee.setValidator(QDoubleValidator(0.0, 100.0, 2))
        self.otomasyon_sonrasi_calisma_gunu = QLineEdit()
        self.otomasyon_sonrasi_calisma_gunu.setValidator(QDoubleValidator(0.0, 366.0, 0))

        # Birim Ürün Fiyatı
        self.birim_urun_fiyati = QLineEdit()
        self.birim_urun_fiyati.setValidator(QDoubleValidator(0.0, 1e12, 2))

        # Mevcut Durum Grubu
        mevcut_layout.addRow("Maksimum Günlük Kapasite (adet/gün):", self.max_kapasite)
        mevcut_layout.addRow("OEE (%) :", self.oee_mevcut)
        mevcut_layout.addRow("Yıllık Çalışma Günü:", self.calisma_gunu)
        mevcut_grup.setLayout(mevcut_layout)

        otomasyon_grup = QGroupBox("Otomasyon Sonrası")
        otomasyon_layout = QFormLayout()
        otomasyon_layout.addRow("Yeni Maksimum Kapasite (adet/gün):", self.otomasyon_sonrasi_max_kapasite)
        otomasyon_layout.addRow("Yeni OEE (%):", self.otomasyon_sonrasi_oee)
        otomasyon_layout.addRow("Yeni Yıllık Çalışma Günü:", self.otomasyon_sonrasi_calisma_gunu)
        otomasyon_grup.setLayout(otomasyon_layout)

        birim_grup = QGroupBox("Birim Ürün Fiyatı (TL/adet)")
        birim_layout = QFormLayout()
        birim_layout.addRow("Ürün Birim Fiyatı (TL):", self.birim_urun_fiyati)
        birim_grup.setLayout(birim_layout)

        ana_layout.addWidget(mevcut_grup)
        ana_layout.addWidget(otomasyon_grup)
        ana_layout.addWidget(birim_grup)

        icon = self.style().standardIcon(QStyle.SP_ArrowForward)
        self.tab_widget.addTab(sayfa, icon, "Verimlilik")

    def kalite_sayfasi(self):
        """Create the tab for quality metrics."""
        sayfa = QWidget()
        ana_layout = QVBoxLayout()
        sayfa.setLayout(ana_layout)

        mevcut_grup = QGroupBox("Mevcut Durum")
        mevcut_layout = QFormLayout()

        # Mevcut Durum Input Alanları
        self.iade_urun_sayisi = QLineEdit()
        self.iade_urun_sayisi.setValidator(QDoubleValidator(0.0, 1e12, 2))
        self.ortalama_iade_maliyeti = QLineEdit()
        self.ortalama_iade_maliyeti.setValidator(QDoubleValidator(0.0, 1e12, 2))
        self.musteri_sikayet_sayisi = QLineEdit()
        self.musteri_sikayet_sayisi.setValidator(QDoubleValidator(0.0, 1e12, 2))
        self.ortalama_sikayet_maliyeti = QLineEdit()
        self.ortalama_sikayet_maliyeti.setValidator(QDoubleValidator(0.0, 1e12, 2))

        # Otomasyon Sonrası Input Alanları
        self.otomasyon_sonrasi_iade_urun_sayisi = QLineEdit()
        self.otomasyon_sonrasi_iade_urun_sayisi.setValidator(QDoubleValidator(0.0, 1e12, 2))
        self.otomasyon_sonrasi_iade_maliyeti = QLineEdit()
        self.otomasyon_sonrasi_iade_maliyeti.setValidator(QDoubleValidator(0.0, 1e12, 2))
        self.otomasyon_sonrasi_sikayet_sayisi = QLineEdit()
        self.otomasyon_sonrasi_sikayet_sayisi.setValidator(QDoubleValidator(0.0, 1e12, 2))
        self.otomasyon_sonrasi_sikayet_maliyeti = QLineEdit()
        self.otomasyon_sonrasi_sikayet_maliyeti.setValidator(QDoubleValidator(0.0, 1e12, 2))

        # Mevcut Durum Grubu
        mevcut_layout.addRow("Yıllık İade Ürün Sayısı:", self.iade_urun_sayisi)
        mevcut_layout.addRow("Ortalama İade Maliyeti:", self.ortalama_iade_maliyeti)
        mevcut_layout.addRow("Yıllık Müşteri Şikayet Sayısı:", self.musteri_sikayet_sayisi)
        mevcut_layout.addRow("Ortalama Şikayet Maliyeti:", self.ortalama_sikayet_maliyeti)
        mevcut_grup.setLayout(mevcut_layout)

        otomasyon_grup = QGroupBox("Otomasyon Sonrası")
        otomasyon_layout = QFormLayout()
        otomasyon_layout.addRow("Yıllık İade Ürün Sayısı:", self.otomasyon_sonrasi_iade_urun_sayisi)
        otomasyon_layout.addRow("Ortalama İade Maliyeti:", self.otomasyon_sonrasi_iade_maliyeti)
        otomasyon_layout.addRow("Yıllık Müşteri Şikayet Sayısı:", self.otomasyon_sonrasi_sikayet_sayisi)
        otomasyon_layout.addRow("Ortalama Şikayet Maliyeti:", self.otomasyon_sonrasi_sikayet_maliyeti)
        otomasyon_grup.setLayout(otomasyon_layout)

        ana_layout.addWidget(mevcut_grup)
        ana_layout.addWidget(otomasyon_grup)

        icon = self.style().standardIcon(QStyle.SP_MessageBoxInformation)
        self.tab_widget.addTab(sayfa, icon, "Kalite")

    def _to_float(self, line_edit: QLineEdit) -> float:
        """Return the numeric value of a QLineEdit or zero.

        Displays a warning and returns 0.0 if conversion fails.
        """
        text = line_edit.text()
        if not text:
            return 0.0
        try:
            return float(text)
        except ValueError:
            QMessageBox.warning(self, "Geçersiz Giriş", f"'{text}' sayısal bir değer değil.")
            return 0.0

    def roi_hesapla(self):
        """Collect user inputs, perform calculations and export the report."""
        try:
            # Tüm input alanlarından verileri al
            sirket_adi = self.sirket_adi.text() or "Bilinmeyen Şirket"
            proje_adi = self.proje_adi.text() or "Isimsiz Proje"

            data = {
                "sirket_adi": sirket_adi,
                "proje_adi": proje_adi,
                "mevcut_isci_sayisi": self._to_float(self.mevcut_isci_sayisi),
                "ortalama_maas": self._to_float(self.ortalama_maas),
                "mevcut_vardiya_sayisi": self._to_float(self.mevcut_vardiya_sayisi),
                "otomasyon_sonrasi_isci_sayisi": self._to_float(self.otomasyon_sonrasi_isci_sayisi),
                "otomasyon_sonrasi_maas": self._to_float(self.otomasyon_sonrasi_maas),
                "otomasyon_sonrasi_vardiya_sayisi": self._to_float(self.otomasyon_sonrasi_vardiya_sayisi),
                "max_kapasite": self._to_float(self.max_kapasite),
                "oee_mevcut": self._to_float(self.oee_mevcut),
                "calisma_gunu": self._to_float(self.calisma_gunu),
                "otomasyon_sonrasi_max_kapasite": self._to_float(self.otomasyon_sonrasi_max_kapasite),
                "otomasyon_sonrasi_oee": self._to_float(self.otomasyon_sonrasi_oee),
                "otomasyon_sonrasi_calisma_gunu": self._to_float(self.otomasyon_sonrasi_calisma_gunu),
                "birim_urun_fiyati": self._to_float(self.birim_urun_fiyati),
                "iade_urun_sayisi": self._to_float(self.iade_urun_sayisi),
                "ortalama_iade_maliyeti": self._to_float(self.ortalama_iade_maliyeti),
                "musteri_sikayet_sayisi": self._to_float(self.musteri_sikayet_sayisi),
                "ortalama_sikayet_maliyeti": self._to_float(self.ortalama_sikayet_maliyeti),
                "otomasyon_sonrasi_iade_urun_sayisi": self._to_float(self.otomasyon_sonrasi_iade_urun_sayisi),
                "otomasyon_sonrasi_iade_maliyeti": self._to_float(self.otomasyon_sonrasi_iade_maliyeti),
                "otomasyon_sonrasi_sikayet_sayisi": self._to_float(self.otomasyon_sonrasi_sikayet_sayisi),
                "otomasyon_sonrasi_sikayet_maliyeti": self._to_float(self.otomasyon_sonrasi_sikayet_maliyeti),
            }

            dosya_adi = f"{sirket_adi}_{proje_adi}_ROI_Raporu_{datetime.now().strftime('%Y%m%d')}.xlsx"

            if self.use_xlsxwriter.isChecked():
                create_xlsxwriter_report(data, dosya_adi)
                QMessageBox.information(
                    self,
                    "Başarılı",
                    f"Gelişmiş ROI Raporu {dosya_adi} olarak kaydedildi!",
                )
                return

            # Excel oluşturma
            wb = openpyxl.Workbook()

            # Tarih bilgisini ekle
            ana_sayfa_data = [
                ["Otomasyon ve Proses İyileştirme ROI Hesaplama Aracı"],
                [],
                ["Şirket Adı:", "", "🏢"],
                ["Proje Adı:", "", "📋"],
                ["Tarih:", datetime.now().strftime("%d.%m.%Y"), "📅"],
                [],
                ["Bu araç, otomasyon ve proses iyileştirme projelerinin finansal etkisini hesaplamak için tasarlanmıştır."],
                [],
                ["İçindekiler:"],
                ["1. Maliyet Tasarrufu Hesaplama", "💰"],
                ["2. Verimlilik Artışı Hesaplama", "📈"],
                ["3. Kalite İyileştirme Hesaplama", "🔍"],
                ["4. Özet ve ROI Analizi", "📊"],
                ["5. NPV ve ROI Bilgi", "ℹ️"]
            ]

            # Sayfa oluşturma ve veri şablonlarını doldurma
            sayfalar = [
                ("Ana Sayfa", ana_sayfa_data),
                ("1-Maliyet Tasarrufu", maliyet_tasarrufu_data),
                ("2-Verimlilik Artışı", verimlilik_artisi_data),
                ("3-Kalite İyileştirme", kalite_iyilestirme_data),
                ("4-Özet ve ROI", ozet_roi_data),
                ("5-NPV_ROI Bilgi", npv_roi_bilgi_data)
            ]

            for baslik, sablon in sayfalar:
                sayfa = wb.create_sheet(baslik)
                for satir_index, satir in enumerate(sablon, start=1):
                    for sutun_index, deger in enumerate(satir, start=1):
                        hucre = sayfa.cell(row=satir_index, column=sutun_index)
                        hucre.value = deger

            # İlk boş sayfayı sil
            wb.remove(wb['Sheet'])

            # Ana Sayfa bilgilerini güncelle
            ana_sayfa = wb["Ana Sayfa"]
            ana_sayfa['B3'].value = sirket_adi
            ana_sayfa['B4'].value = proje_adi

            # Verileri doldur
            maliyet_sayfasi = wb["1-Maliyet Tasarrufu"]
            maliyet_sayfasi['B5'].value = self._to_float(self.mevcut_isci_sayisi)
            maliyet_sayfasi['B6'].value = self._to_float(self.ortalama_maas)
            maliyet_sayfasi['B7'].value = self._to_float(self.mevcut_vardiya_sayisi)

            maliyet_sayfasi['B11'].value = self._to_float(self.otomasyon_sonrasi_isci_sayisi)
            maliyet_sayfasi['B12'].value = self._to_float(self.otomasyon_sonrasi_maas)
            maliyet_sayfasi['B13'].value = self._to_float(self.otomasyon_sonrasi_vardiya_sayisi)

            # Verimlilik Sayfası Verileri
            verimlilik_sayfasi = wb["2-Verimlilik Artışı"]
            verimlilik_sayfasi['B5'].value = self._to_float(self.max_kapasite)
            verimlilik_sayfasi['B6'].value = self._to_float(self.oee_mevcut) / 100
            verimlilik_sayfasi['B7'].value = self._to_float(self.calisma_gunu)

            verimlilik_sayfasi['B10'].value = self._to_float(self.otomasyon_sonrasi_max_kapasite)
            verimlilik_sayfasi['B11'].value = self._to_float(self.otomasyon_sonrasi_oee) / 100
            verimlilik_sayfasi['B12'].value = self._to_float(self.otomasyon_sonrasi_calisma_gunu)

            verimlilik_sayfasi['B15'].value = self._to_float(self.birim_urun_fiyati)

            # Kalite sayfası verileri
            kalite_sayfasi = wb["3-Kalite İyileştirme"]
            kalite_sayfasi['B5'].value = self._to_float(self.iade_urun_sayisi)
            kalite_sayfasi['B6'].value = self._to_float(self.ortalama_iade_maliyeti)
            kalite_sayfasi['B7'].value = self._to_float(self.musteri_sikayet_sayisi)
            kalite_sayfasi['B8'].value = self._to_float(self.ortalama_sikayet_maliyeti)

            kalite_sayfasi['B12'].value = self._to_float(self.otomasyon_sonrasi_iade_urun_sayisi)
            kalite_sayfasi['B13'].value = self._to_float(self.otomasyon_sonrasi_iade_maliyeti)
            kalite_sayfasi['B14'].value = self._to_float(self.otomasyon_sonrasi_sikayet_sayisi)
            kalite_sayfasi['B15'].value = self._to_float(self.otomasyon_sonrasi_sikayet_maliyeti)

            # Özet ve ROI sayfasını tanımla
            ozet_sayfasi = wb["4-Özet ve ROI"]

            # Hesaplama fonksiyonlarını çağır
            maliyet_tasarrufu_hesapla(maliyet_sayfasi)
            verimlilik_artisi_hesapla(verimlilik_sayfasi)
            kalite_iyilestirme_hesapla(kalite_sayfasi)

            # Özet ve ROI sayfasına GUI’den gelen değerleri taşı
            ozet_sayfasi['B15'].value = maliyet_sayfasi['B14'].value  # Yıllık Maliyet Tasarrufu
            ozet_sayfasi['B16'].value = verimlilik_sayfasi['B20'].value  # Yıllık Verimlilik Artışı
            ozet_sayfasi['B17'].value = kalite_sayfasi['B19'].value  # Yıllık Kalite İyileştirme

            # B sütunundaki formülleri ekle
            ozet_sayfasi['B18'].value = '=SUM(B15:B17)'  # Toplam Yıllık Getiri
            ozet_sayfasi['B21'].value = '=B11'  # Toplam Yatırım Maliyeti (Finansal Analiz)
            ozet_sayfasi['B22'].value = '=IF(B11>0,B18/B11,0)'  # ROI (%)
            ozet_sayfasi['B23'].value = '=IF(B18>0,B11/B18,0)'  # Geri Ödeme Süresi

            # Yardımcı sütunlar (D, E, F) ekle
            ozet_sayfasi['D1'].value = "Yıl"
            ozet_sayfasi['E1'].value = "Yıllık Getiri (TL)"
            ozet_sayfasi['F1'].value = "Bugünkü Değer (TL)"
            for i in range(1, 6):
                cell = ozet_sayfasi[f'D{i+3}']
                cell.value = i  # D4:D8: 1’den 5’e kadar yıllar
                cell.number_format = '0 "Yıl"'

            # E ve F sütunlarındaki formülleri ekle
            ozet_sayfasi['E4'].value = '=B18'  # 1. yıl getirisi
            ozet_sayfasi['E5'].value = '=E4*(1+B34)'  # 2. yıl getirisi (Yıllık İşçilik Maaş Artışı)
            ozet_sayfasi['E6'].value = '=E5*(1+B34)'  # 3. yıl getirisi
            ozet_sayfasi['E7'].value = '=E6*(1+B34)'  # 4. yıl getirisi
            ozet_sayfasi['E8'].value = '=E7*(1+B34)'  # 5. yıl getirisi

            ozet_sayfasi['F4'].value = '=E5/(1+B32)^1'  # 1. yıl bugünkü değeri
            ozet_sayfasi['F5'].value = '=E6/(1+B32)^2'  # 2. yıl bugünkü değeri
            ozet_sayfasi['F6'].value = '=E7/(1+B32)^3'  # 3. yıl bugünkü değeri
            ozet_sayfasi['F7'].value = '=E8/(1+B32)^4'  # 4. yıl bugünkü değeri
            ozet_sayfasi['F8'].value = '=E9/(1+B32)^5'  # 5. yıl bugünkü değeri

            # Kümülatif getiri ve toplam yatırım sütunları için formüller
            ozet_sayfasi['G4'].value = '=E4'
            ozet_sayfasi['G5'].value = '=G4+E5'
            ozet_sayfasi['G6'].value = '=G5+E6'
            ozet_sayfasi['G7'].value = '=G6+E7'
            ozet_sayfasi['G8'].value = '=G7+E8'
            for r in range(4, 9):
                ozet_sayfasi[f'H{r}'].value = f'=B11*(1+B32)^{r-3}'
                ozet_sayfasi[f'I{r}'].value = f'=H{r}/(1+B32)^{r-3}'

            # NPV ve Banka Faizi formüllerini ekle
            ozet_sayfasi['B24'].value = '=SUM(F4:INDEX(F4:F8,B33))-B11'  # NPV
            ozet_sayfasi['B27'].value = '=IF(B11>0,B11*(1+B32)^B33,0)'  # Banka Faizi ile Elde Edilecek Getiri
            ozet_sayfasi['B28'].value = '=SUM(OFFSET(I3,1,0,B33,1))-B11'

            # Grafik ekle
            grafik_ekle(wb)

            # Sayfa biçimlendirmeleri
            for sayfa in wb.sheetnames:
                sutun_genislikleri_ayarla(wb[sayfa])
                sayfa_bicimlendir(wb[sayfa])

            # Dosya kaydetme
            dosya_adi = f"{sirket_adi}_{proje_adi}_ROI_Raporu_{datetime.now().strftime('%Y%m%d')}.xlsx"
            wb.save(dosya_adi)

            # Kullanıcıya bilgilendirme
            QMessageBox.information(
                self, 
                "Başarılı", 
                f"ROI Raporu {dosya_adi} olarak kaydedildi!"
            )
        except Exception as e:
            QMessageBox.warning(self, "Hata", f"Hesaplamada sorun oluştu: {str(e)}")

def sayfa_bicimlendir(sheet):
    """Apply openpyxl styles to the provided worksheet."""
    styles = get_common_styles()

    # Başlık ve alt başlık stilleri
    baslik_font = Font(name=styles['font_name'], size=14, bold=True, color="FFFFFF")
    alt_baslik_font = Font(name=styles['font_name'], size=12, bold=True)
    normal_font = Font(name=styles['font_name'], size=11)
    bold_font = Font(name=styles['font_name'], size=11, bold=True)  # Sonuç için kalın font

    # Dolgular
    baslik_dolgu = PatternFill(start_color=styles['header_bg'], end_color=styles['header_bg'], fill_type="solid")
    alt_baslik_dolgu = PatternFill(start_color=styles['subheader_bg'], end_color=styles['subheader_bg'], fill_type="solid")

    # Kenarlık
    kenarlik = Border(
        left=Side(style='thin', color=styles['border_color']),
        right=Side(style='thin', color=styles['border_color']),
        top=Side(style='thin', color=styles['border_color']),
        bottom=Side(style='thin', color=styles['border_color'])
    )

    # Hücreleri biçimlendir
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = normal_font
            cell.border = kenarlik
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # İlk satırı başlık olarak biçimlendir
            if cell.row == 1:
                cell.font = baslik_font
                cell.fill = baslik_dolgu

            # İkinci satırı alt başlık olarak biçimlendir
            if cell.row == 2:
                cell.font = alt_baslik_font
                cell.fill = alt_baslik_dolgu

            # NPV ve ROI Bilgi sayfasında "Sonuç:" satırlarını kalın yap
            if sheet.title == "5-NPV_ROI Bilgi" and cell.value == "Sonuç:":
                cell.font = bold_font

def sutun_genislikleri_ayarla(sheet):
    """Automatically adjust column widths on an openpyxl worksheet."""
    # Sütun genişliklerini otomatik ayarla, ancak bazı sütunlar için sabit genişlik belirle
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        # Özet sheet’inde B sütunu için genişliği 32’ye sabitle
        if sheet.title == "4-Özet ve ROI" and column_letter == "B":
            sheet.column_dimensions[column_letter].width = 32
            continue

        # NPV_ROI Bilgi sheet’inde A sütunu için genişliği 50’ye sabitle
        if sheet.title == "5-NPV_ROI Bilgi" and column_letter == "A":
            sheet.column_dimensions[column_letter].width = 50
            continue

        # Diğer sütunlar için otomatik genişlik ayarlama
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            except Exception as e:
                logging.exception(
                    f"Error processing cell {cell.coordinate} in '{sheet.title}': {e}"
                )

        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

def grafik_ekle(wb):
    """Insert default charts into the summary worksheet."""
    sheet = wb["4-Özet ve ROI"]

    # Yatırım Maliyetleri Grafiği
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 12  # Daha modern bir stil
    chart1.title = "Proje Yatırım Maliyetleri"
    chart1.height = 15
    chart1.width = 20
    chart1.titleOverlay = False  # Başlık grafiğin üstüne taşınır
    chart1.dLbls = DataLabelList()
    chart1.dLbls.showVal = True

    data = Reference(sheet, min_col=2, min_row=4, max_row=8)
    cats = Reference(sheet, min_col=1, min_row=4, max_row=8)

    chart1.add_data(data, titles_from_data=False)
    chart1.set_categories(cats)

    sheet.add_chart(chart1, "H1")  # Grafiği H1’e taşı

    # Getiriler Grafiği
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 12  # Daha modern bir stil
    chart2.title = "Yıllık Getiriler"
    chart2.height = 15
    chart2.width = 20
    chart2.titleOverlay = False  # Başlık grafiğin üstüne taşınır
    chart2.dLbls = DataLabelList()
    chart2.dLbls.showVal = True

    data = Reference(sheet, min_col=2, min_row=14, max_row=16)
    cats = Reference(sheet, min_col=1, min_row=14, max_row=16)

    chart2.add_data(data, titles_from_data=False)
    chart2.set_categories(cats)

    sheet.add_chart(chart2, "H21")  # Grafiği H21’ye taşı

def main():
    """Launch the ROI calculator GUI."""
    app = QApplication(sys.argv)
    pencere = ROIHesaplamaArayuzu()
    pencere.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
